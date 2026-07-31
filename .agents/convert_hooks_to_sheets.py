import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

md_path = os.path.join(os.path.dirname(__file__), "HOOKS_GUIDE.md")
excel_path = os.path.join(os.path.dirname(__file__), "HOOKS_GUIDE.xlsx")

with open(md_path, "r", encoding="utf-8") as f:
    content = f.read()

# Create Workbook
wb = openpyxl.Workbook()
wb.remove(wb.active) # Remove default sheet

# Helper for styling
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
regular_font = Font(name="Segoe UI", size=10)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Parse Summary Table
hooks_data = []
for line in content.split('\n'):
    if line.strip().startswith('| **'):
        parts = [p.strip() for p in line.split('|')[1:-1]]
        if len(parts) >= 5:
            # Clean formatting
            name = parts[0].replace('**', '')
            agents = parts[1].replace('`', '')
            model_level = parts[2].replace('<br>', '\n')
            triggers = parts[3].replace('"', '').replace('`', '')
            goal = parts[4]
            
            # Extract ID from name
            m = re.match(r'(\d+)\.', name)
            idx = int(m.group(1)) if m else len(hooks_data) + 1
            
            hooks_data.append({
                "id": idx,
                "name": name,
                "agents": agents,
                "model_level": model_level,
                "triggers": triggers,
                "goal": goal
            })

# Parse Detailed Sections for full text
detail_blocks = re.findall(r'###\s*(\d+)\.\s*(.*?)\n(.*?)(?=(?:###\s*\d+|\Z))', content, re.DOTALL)
details_dict = {}
for num_str, title, block_text in detail_blocks:
    num = int(num_str)
    details_dict[num] = block_text.strip()

# Add detail text to hooks_data
for hook in hooks_data:
    hook["details"] = details_dict.get(hook["id"], "")

headers = ["م", "اسم الخطاف", "المحفزات (الخطاف للنسخ)", "النموذج الأساسي والبديل", "الوكلاء المسؤولون", "الهدف الأساسي", "الوصف والتفاصيل"]
column_widths = {1: 6, 2: 28, 3: 40, 4: 45, 5: 35, 6: 45, 7: 80}

def format_sheet(ws, data_rows):
    ws.views.sheetView[0].rightToLeft = True
    ws.append(headers)
    
    # Format Headers
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    # Append Data
    for row_data in data_rows:
        ws.append(row_data)
        
    # Format Data Rows
    for row in ws.iter_rows(min_row=2, max_row=len(data_rows)+1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", horizontal="right", wrap_text=True)
            
    # Set Column Widths
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

# 1. Summary Sheet (All Hooks)
ws_summary = wb.create_sheet(title="الملخص العام")
all_rows = []
for hook in hooks_data:
    all_rows.append([hook["id"], hook["name"], hook["triggers"], hook["model_level"], hook["agents"], hook["goal"], hook["details"]])
format_sheet(ws_summary, all_rows)

# 2. Individual Sheets
for hook in hooks_data:
    clean_name = re.sub(r'^\d+\.\s*', '', hook['name'])
    if len(clean_name) > 25:
        clean_name = clean_name[:25]
    sheet_title = f"{hook['id']}. {clean_name}"
    
    ws = wb.create_sheet(title=sheet_title)
    single_row = [[hook["id"], hook["name"], hook["triggers"], hook["model_level"], hook["agents"], hook["goal"], hook["details"]]]
    format_sheet(ws, single_row)

# Save workbook
wb.save(excel_path)
print("SUCCESS: Excel workbook updated with Triggers in copyable column and independent Details!")

